from __future__ import annotations

import re
import unittest
from contextlib import ExitStack
from unittest.mock import patch

import app as app_module
import plot_functions
import site_routes


LEGACY_PLOT_ROUTES = ["/plot_bar", "/plot_box", "/plot_hist"]
REMOVED_PLACEHOLDER_ROUTES = [
    "/plots/",
    "/plots/bar",
    "/plots/box",
    "/plots/hist",
    "/plots/scatter",
    "/plots/stats",
]


class VisualisationRouteTests(unittest.TestCase):
    def setUp(self):
        app_module.app.config.update(TESTING=True, LOGIN_DISABLED=True)
        self.client = app_module.app.test_client()
        self.patches = ExitStack()
        self.patches.enter_context(patch.object(site_routes, "_current_email", return_value="user@example.com"))
        self.patches.enter_context(patch.object(site_routes, "get_owned_sites", return_value=[]))
        self.get_rows = self.patches.enter_context(patch.object(site_routes, "get_owned_sites_rows", return_value=[]))
        plot_functions._REF_DF = None

    def tearDown(self):
        self.patches.close()
        app_module.app.config["LOGIN_DISABLED"] = False
        plot_functions._REF_DF = None

    def test_database_dropdown_links_to_site_database_and_workbench(self):
        response = self.client.get("/")
        page = response.get_data(as_text=True)
        database_menu = re.search(
            r"Database &#9662;.*?<ul class=\"dropdown\">(.*?)</ul>",
            page,
            flags=re.DOTALL,
        ).group(1)
        hrefs = re.findall(r'href="([^"]+)"', database_menu)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(hrefs, ["/sites", "/data_analysis"])
        self.assertIn(">Site Database</a>", database_menu)
        self.assertIn(">Data Workbench</a>", database_menu)
        self.assertLess(page.index(">Home</a>"), page.index("Database &#9662;"))
        self.assertLess(page.index("Database &#9662;"), page.index("Analytical Model &#9662;"))

    def test_database_page_does_not_include_workbench_section(self):
        response = self.client.get("/sites")
        page = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("<h2>Site Database</h2>", page)
        self.assertNotIn("Visual Analysis", page)
        self.assertNotIn('class="primary-btn" href="/data_analysis"', page)

    def test_reference_database_uses_legacy_data_and_renders_browsing_controls(self):
        headers, rows = site_routes._reference_database_rows()

        self.assertEqual(len(headers), 16)
        self.assertEqual(len(rows), 112)
        self.assertEqual(rows[0]["Site Unit"], "Hill AFB,UT,site 870")

        response = self.client.get("/sites")
        page = response.get_data(as_text=True)
        reference_section = page.split('<h2 class="site-user-database-heading">', 1)[0]

        self.assertEqual(response.status_code, 200)
        self.assertEqual(page.count("data-reference-row"), 112)
        self.assertIn("data-database-size", reference_section)
        self.assertIn("data-database-search", reference_section)
        self.assertIn("data-database-sort", reference_section)
        self.assertIn("data-database-pagination", reference_section)
        self.assertIn("<h3>Reference Database</h3>", reference_section)
        self.assertNotIn("Browse the built-in site dataset", page)
        self.assertNotIn("data-database-export", reference_section)

    def test_site_database_places_reference_and_upload_download_actions(self):
        response = self.client.get("/sites")
        page = response.get_data(as_text=True)
        reference_section, user_section = page.split(
            '<h2 class="site-user-database-heading">', 1
        )
        upload_form = re.search(
            r'<form method="POST" enctype="multipart/form-data" class="site-form csv-form">(.*?)</form>',
            user_section,
            flags=re.DOTALL,
        ).group(1)

        self.assertEqual(response.status_code, 200)
        self.assertIn('href="/sites/reference.csv" download', reference_section)
        self.assertIn('href="/dispersivity-data"', reference_section)
        self.assertIn("Download Reference Database", reference_section)
        self.assertIn("Dispersivity Data", reference_section)
        self.assertIn("Upload CSV", upload_form)
        self.assertIn("Download Sample File", upload_form)
        self.assertLess(upload_form.index("Upload CSV"), upload_form.index("Download Sample File"))

    def test_dispersivity_page_uses_legacy_data_table_and_plots(self):
        headers, rows = site_routes._dispersivity_database_rows()

        self.assertEqual(headers, ["Reference", "Alpha_T", "Alpha_TV"])
        self.assertEqual(len(rows), 150)
        self.assertEqual(rows[0]["Reference"], "Aksoy and Guney (2010)")

        response = self.client.get("/dispersivity-data")
        page = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("<h2>Dispersivity Data</h2>", page)
        self.assertEqual(page.count("data-dispersivity-row"), 150)
        self.assertIn("data-database-size", page)
        self.assertIn("data-database-search", page)
        self.assertIn("data-database-sort", page)
        self.assertIn("data-database-pagination", page)
        self.assertIn('href="/static/fig1_plots.csv" download', page)
        for image in ("ticks.png", "box.png", "scatter.png"):
            self.assertIn(f'/static/DispersivityPlots/{image}', page)
        self.assertEqual(self.client.get("/static/fig1_plots.csv").status_code, 200)

    def test_reference_database_exports_filtered_csv_excel_and_pdf(self):
        query = "?q=Madison%20ANGB&sort=0&dir=desc"

        csv_response = self.client.get("/sites/reference.csv" + query)
        self.assertEqual(csv_response.status_code, 200)
        self.assertEqual(csv_response.mimetype, "text/csv")
        csv_lines = csv_response.data.decode("utf-8-sig").splitlines()
        self.assertEqual(len(csv_lines), 2)
        self.assertIn("Madison ANGB,WI", csv_lines[1])

        excel_response = self.client.get("/sites/reference.xlsx" + query)
        self.assertEqual(excel_response.status_code, 200)
        self.assertEqual(
            excel_response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(excel_response.data.startswith(b"PK"))
        from io import BytesIO
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(excel_response.data), read_only=True)
        excel_rows = list(workbook.active.values)
        workbook.close()
        self.assertEqual(len(excel_rows), 2)
        self.assertEqual(excel_rows[1][1], "Madison ANGB,WI")

        pdf_response = self.client.get("/sites/reference.pdf" + query)
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response.mimetype, "application/pdf")
        self.assertTrue(pdf_response.data.startswith(b"%PDF"))

        self.assertEqual(self.client.get("/sites/reference.zip").status_code, 404)

    def test_user_database_controls_and_exports_current_rows(self):
        self.get_rows.return_value = [
            {
                "id": 42,
                "site_unit": "Alpha Site",
                "compound": "Benzene",
                "plume_length": 12.5,
                "extra_data": {"Project": "North"},
            },
            {
                "id": 43,
                "site_unit": "Beta Site",
                "compound": "Toluene",
                "plume_length": 8.0,
                "extra_data": {"Project": "South"},
            },
        ]

        response = self.client.get("/sites")
        page = response.get_data(as_text=True)
        user_section = page.split('<h2 class="site-user-database-heading">', 1)[1]

        self.assertEqual(response.status_code, 200)
        self.assertIn("Download Sample File", user_section)
        self.assertIn('/static/sample_db.csv', user_section)
        self.assertNotIn("upload-notation-list", user_section)
        self.assertIn("data-database-size", user_section)
        self.assertIn("data-database-search", user_section)
        self.assertIn("data-database-pagination", user_section)
        for action in ("copy", "csv", "excel", "pdf", "print"):
            self.assertIn(f'data-database-export="{action}"', user_section)
        self.assertIn('data-database-csv-url="/sites/user.csv"', user_section)
        self.assertIn('data-database-excel-url="/sites/user.xlsx"', user_section)
        self.assertIn('data-database-pdf-url="/sites/user.pdf"', user_section)

        query = "?q=Alpha%20Site"
        csv_response = self.client.get("/sites/user.csv" + query)
        self.assertEqual(csv_response.status_code, 200)
        csv_lines = csv_response.data.decode("utf-8-sig").splitlines()
        self.assertEqual(len(csv_lines), 2)
        self.assertIn("Alpha Site", csv_lines[1])
        self.assertNotIn("Beta Site", csv_response.data.decode("utf-8-sig"))

        excel_response = self.client.get("/sites/user.xlsx" + query)
        self.assertEqual(excel_response.status_code, 200)
        from io import BytesIO
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(excel_response.data))
        excel_rows = list(workbook.active.values)
        worksheet_title = workbook.active.title
        self.assertIsNone(workbook.active.freeze_panes)
        self.assertIsNone(workbook.active.auto_filter.ref)
        workbook.close()
        self.assertEqual(worksheet_title, "Site Database")
        self.assertEqual(excel_rows[0][0], "Site Database")
        self.assertEqual(excel_rows[1][:3], ("ID", "Site unit", "Compound"))
        self.assertEqual(excel_rows[2][:3], (1, "Alpha Site", "Benzene"))
        self.assertEqual(len(excel_rows), 3)

        pdf_response = self.client.get("/sites/user.pdf" + query)
        self.assertEqual(pdf_response.status_code, 200)
        self.assertTrue(pdf_response.data.startswith(b"%PDF"))
        self.assertEqual(self.client.get("/sites/user.zip").status_code, 404)

    def test_database_page_hides_legacy_filter_and_sort_tools(self):
        self.get_rows.return_value = [{"id": 1, "site_unit": "A", "compound": "B", "extra_data": {}}]
        response = self.client.get("/sites")
        page = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("Clear Database", page)
        self.assertNotIn("Filter Columns", page)
        self.assertNotIn("Sort Rows", page)
        self.assertNotIn("Apply Filters", page)

    def test_remaining_routes_render_bokeh_without_reference_csv(self):
        with patch.object(plot_functions.Path, "exists", return_value=False):
            for path in LEGACY_PLOT_ROUTES:
                with self.subTest(path=path):
                    response = self.client.get(path)
                    page = response.get_data(as_text=True)

                    self.assertEqual(response.status_code, 200)
                    self.assertIn('src="/static/extensions/panel/', page)
                    self.assertIn("cdn.bokeh.org/bokeh/release/bokeh-", page)
                    self.assertIn("Bokeh.embed.embed_items", page)
                    self.assertIn("No data available", page)
                    self.assertIn('class="card-like analysis-plot-shell bokeh-output"', page)
                    self.assertIn('"sizing_mode":"stretch_width"', page)

    def test_data_analysis_wrapper_renders_panel_iframe(self):
        response = self.client.get("/data_analysis")
        page = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("/panel_data_analysis", page)
        self.assertIn('class="panel-frame output-panel-frame data-analysis-frame"', page)
        self.assertNotIn("Back to Site Database", page)

    def test_data_analysis_wrapper_passes_an_owned_completed_aem_job(self):
        meta = {"email": "user@example.com", "kind": "aem_forward"}
        with (
            patch.object(site_routes, "load_job_meta", return_value=meta),
            patch.object(site_routes, "job_status", return_value={"status": "done"}),
        ):
            response = self.client.get("/data_analysis?aem_job=job-1")

        self.assertEqual(response.status_code, 200)
        self.assertIn("/panel_data_analysis?aem_job=job-1", response.get_data(as_text=True))

    def test_data_analysis_wrapper_rejects_another_users_aem_job(self):
        meta = {"email": "other@example.com", "kind": "aem_forward"}
        with patch.object(site_routes, "load_job_meta", return_value=meta):
            response = self.client.get("/data_analysis?aem_job=job-1")

        self.assertEqual(response.status_code, 404)

    def test_placeholder_plot_blueprint_routes_are_removed(self):
        for path in REMOVED_PLACEHOLDER_ROUTES:
            with self.subTest(path=path):
                self.assertEqual(self.client.get(path).status_code, 404)


if __name__ == "__main__":
    unittest.main()
